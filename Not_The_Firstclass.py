import openpyxl
import re
import numpy as np
from sys import *
from openpyxl.styles import PatternFill

Path3 = input('Enter the path of the previous Excel File: ')
Name_Excel = input('Enter the name of this excel file: ')
wb1 = openpyxl.load_workbook(str(Path3) + '\\' + str(Name_Excel) + '.xlsx')
sheet1 = wb1.active
lecture_no = int(input('Enter the lecture number: '))
class_name1 = sheet1.title
n1 = int(sheet1.max_row)
n = n1 - 1
print('Class Name: ', class_name1)

arr = np.arange(100)
roll_object = open('C:\\Users\\Admin\\Documents\\Python practice\\Attendance\\Sample Attendance2.txt')
roll_list = roll_object.readlines()
print(f'Total number of students in {class_name1} are {n}')
total = np.arange(n)
i = 0
while i < len(roll_list):
    roll_object1 = re.compile(r'\d\d\d\d')
    mo = roll_object1.search(roll_list[i])
    for j in range(1, 101):
        arr[i] = mo.group()
    i = i + 1
absent_arr = np.arange(100)
roll1 = sheet1['A2'].value
print('The first roll number of the class is: ', roll1)
for t in range(0, n):
    total[t] = roll1 + t
print('The absent students are: ')
r = 0
count1 = 0
while r < len(total):
    a = 0
    count = 0
    while a < n*3:
        if total[r] == arr[a]:
            count = count + 1
            a = a + 1
        else:
            a = a + 1
    if count == 0:
        absent_arr[count1] = total[r]
        print(total[r])
        count1 = count1 + 1
    r = r + 1
print(f'The total number of absent students are: {count1}')
print('Absent array: ', absent_arr)
green_Fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
red_Fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
for i in range(2, sheet1.max_row + 1):
    if total[i-2] in absent_arr:
        sheet1.cell(row=i, column=lecture_no+2).value = 0
        sheet1.cell(row=i, column=lecture_no+2).fill = red_Fill
    else:
        sheet1.cell(row=i, column=lecture_no+2).value = 1
        sheet1.cell(row=i, column=lecture_no+2).fill = green_Fill
Name1 = input('Enter the name of the new excel file: ')
Path3 = input('Enter the path of the location where the excel file should be saved: ')

wb1.save(str(Path3) + '\\' + str(Name1) + '.xlsx')
print('A new Excel file is created Successfully')
wb1.save(str(Path3) + '\\' + str(Name1) + '.xlsx')
exit()
