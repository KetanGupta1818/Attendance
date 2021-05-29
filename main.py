# import re
import openpyxl
# import numpy as np
from openpyxl.styles import Font, Alignment, PatternFill
from NameFinding import *

response1 = int(input('If this is first class then Enter 1 \n'
                      'Or Enter 2 '))
if response1 == 2:
    from Not_The_Firstclass import *
n = int(input('Enter the total number of students: '))
class_name = input('Enter the name of class: ')
total_lectures = int(input('Enter the total number of lectures to be taken: '))
arr = np.arange(n*4)  # Taking into consideration, Maximum messages by a single student during attendance time is 2
roll1 = int(input('Enter the roll number of first student: '))

roll_object = open(str(Path1) + '\\' + str(Name_of_Text) + '.txt')
roll_list = roll_object.readlines()

total = np.arange(n)
i = 0
while i < len(roll_list):
    roll_object1 = re.compile(r'\d\d\d\d')
    mo = roll_object1.search(roll_list[i])
    if mo is not None:
        arr[i] = mo.group()
    i = i + 1

for t in range(0, n):
    total[t] = roll1 + t
print('The absent students are: ')
arr_absent = np.arange(n)
r = 0
count1 = 0
while r < len(total):
    a = 0
    count = 0
    while a < n*4:                  # total is an array of sorted total roll numbers
        if total[r] == arr[a]:      # arr is an array of unsorted roll numbers of present students
            count = count + 1
            a = a + 1
        else:
            a = a + 1
    if count == 0:
        arr_absent[count1] = total[r]
        print(total[r])
        count1 = count1 + 1
    r = r + 1
print(f'The total number of absent students are: {count1}')
print('The array of absent roll no. is: ', arr_absent)
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = class_name
for row_cell in range(2, n + 2):
    column_cell = 1
    sheet.cell(row=row_cell, column=column_cell).value = total[row_cell - 2]
sheet.column_dimensions['A'].width = 18.57
font_object = Font(name='Times New Roman', size=14, bold=True)
sheet['A1'].font = font_object
sheet['A1'] = 'Roll Numbers'
for i in range(2, n + 2):
    cell = sheet.cell(row=i, column=1)
    cell.alignment = Alignment(horizontal='center')

sheet.column_dimensions['B'].width = 24
sheet['B1'].font = font_object
sheet['B1'] = 'Names'
for col_num in range(3, total_lectures + 3):
    val = col_num - 2
    sheet.cell(row=1, column=col_num).value = 'Lecture' + str(val)
yellowFill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
pinkFill = PatternFill(start_color='00FF00FF', end_color='00FF00FF', fill_type='solid')
dark_green_Fill = PatternFill(start_color='00808000', end_color='00808000', fill_type='solid')
blueFill = PatternFill(start_color='0000FFFF', end_color='0000FFFF', fill_type='solid')
sheet['A1'].fill = yellowFill
sheet['B1'].fill = blueFill

# Name Filling on sheets
list_keys = list(dictionary.keys())
list_values = list(dictionary.values())
print(list_values)
print(list_keys)
green_Fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
red_Fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
for i in range(2, sheet.max_row + 1):  # max_row = 26
    Total_Roll = sheet.cell(row=i, column=1).value
    if Total_Roll in list_keys:
        sheet.cell(row=i, column=2).value = dictionary[Total_Roll]
        sheet.cell(row=i, column=3).value = 1
        sheet.cell(row=i, column=3).fill = green_Fill
    else:
        sheet.cell(row=i, column=2).value = '------'
        sheet.cell(row=i, column=3).value = 0
        sheet.cell(row=i, column=3).fill = red_Fill
Name = input('Enter the name of the new excel file: ')
Path2 = input('Enter the path of the location where the excel file should be saved: ')

wb.save(str(Path2) + '\\' + str(Name) + '.xlsx')
print('A new Excel file is created Successfully')
